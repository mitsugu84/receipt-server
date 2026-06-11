import os
import json
import base64
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from flask import Flask, render_template, request, send_file, redirect, url_for, flash, session
from werkzeug.utils import secure_filename
from openai import OpenAI
from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from PIL import Image
from pillow_heif import register_heif_opener


register_heif_opener()

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "change-this-secret-key")

UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
SESSION_DIR = Path("sessions")

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
SESSION_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {
    "jpg",
    "jpeg",
    "png",
    "webp",
    "heic",
    "heif",
}

MAX_FILES_PER_UPLOAD = 10

OPENAI_INPUT_PRICE_PER_1M = float(os.environ.get("OPENAI_INPUT_PRICE_PER_1M", "0.15"))
OPENAI_OUTPUT_PRICE_PER_1M = float(os.environ.get("OPENAI_OUTPUT_PRICE_PER_1M", "0.60"))
USD_JPY_RATE = float(os.environ.get("USD_JPY_RATE", "150"))

client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_SECRET_KEY = os.environ.get("SUPABASE_SECRET_KEY")

supabase = None
if SUPABASE_URL and SUPABASE_SECRET_KEY:
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_SECRET_KEY)
        print("Supabase connected")
    except Exception as e:
        print("Supabase connection error:", e)


def get_session_id() -> str:
    if "receipt_session_id" not in session:
        session["receipt_session_id"] = uuid4().hex
    return session["receipt_session_id"]


def get_session_file_path() -> Path:
    return SESSION_DIR / f"{get_session_id()}.json"


def load_session_records() -> list[dict]:
    path = get_session_file_path()

    if not path.exists():
        return []

    try:
        with open(path, "r", encoding="utf-8") as f:
            records = json.load(f)
        return records if isinstance(records, list) else []
    except Exception as e:
        print("Session load error:", e)
        return []


def save_session_records(records: list[dict]):
    path = get_session_file_path()

    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print("Session save error:", e)


def clear_session_records():
    path = get_session_file_path()

    try:
        if path.exists():
            path.unlink()
    except Exception as e:
        print("Session clear error:", e)


def estimate_cost_usd(prompt_tokens: int, completion_tokens: int) -> float:
    input_cost = prompt_tokens / 1_000_000 * OPENAI_INPUT_PRICE_PER_1M
    output_cost = completion_tokens / 1_000_000 * OPENAI_OUTPUT_PRICE_PER_1M
    return input_cost + output_cost


def estimate_cost_jpy(prompt_tokens: int, completion_tokens: int) -> float:
    return estimate_cost_usd(prompt_tokens, completion_tokens) * USD_JPY_RATE


def save_usage_log(filename: str, usage):
    prompt_tokens = getattr(usage, "prompt_tokens", 0) or 0
    completion_tokens = getattr(usage, "completion_tokens", 0) or 0
    total_tokens = getattr(usage, "total_tokens", 0) or 0

    estimated_cost_yen = round(
        estimate_cost_jpy(prompt_tokens, completion_tokens),
        2
    )

    ip_address = request.headers.get("X-Forwarded-For", request.remote_addr)

    if not supabase:
        print("Supabase is not configured")
        return

    try:
        supabase.table("usage_logs").insert({
            "filename": filename,
            "prompt_tokens": prompt_tokens,
            "completion_tokens": completion_tokens,
            "total_tokens": total_tokens,
            "estimated_cost_yen": estimated_cost_yen,
            "ip_address": ip_address,
        }).execute()

        print("Saved to Supabase")

    except Exception as e:
        print("Supabase save error:", e)


def allowed_file(filename: str) -> bool:
    if "." not in filename:
        return False

    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_EXTENSIONS


def delete_file_safely(path: Path):
    try:
        if path and path.exists():
            path.unlink()
            print(f"Deleted temp file: {path.name}")
    except Exception as e:
        print(f"Delete file error: {path.name}", e)


def convert_heic_to_jpeg_if_needed(image_path: Path) -> Path:
    ext = image_path.suffix.lower()

    if ext not in [".heic", ".heif"]:
        return image_path

    converted_path = image_path.with_suffix(".jpg")

    with Image.open(image_path) as image:
        image = image.convert("RGB")
        image.save(
            converted_path,
            "JPEG",
            quality=95,
            optimize=True
        )

    print(f"HEIC converted -> {converted_path.name}")

    return converted_path


def image_to_data_url(image_path: Path) -> str:
    with open(image_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")

    return f"data:image/jpeg;base64,{b64}"


def analyze_receipt(image_path: Path) -> dict:
    data_url = image_to_data_url(image_path)

    prompt = """
あなたは日本の領収書・レシートを読み取る経費整理アシスタントです。

画像から以下を抽出してください。
必ずJSONだけで返してください。説明文は禁止です。

{
  "date": "YYYY-MM-DD または 不明",
  "shop": "店名 または 不明",
  "amount": 数値のみ。税込合計金額。不明なら0,
  "category": "食費/交通費/工具費/消耗品費/通信費/交際費/医療費/その他 のどれか",
  "memo": "補足。なければ空文字"
}

ルール:

【金額】
- 金額は必ずレシート全体を見て、最終的な支払合計を採用してください。
- 商品単価や小計ではなく、合計、合計(税込)、総合計、お支払額、現計、領収額、領収額合計、クレジット支払額、負担金を優先してください。
- レシート下部に合計金額がある場合は、そこを最優先してください。
- お預り金、お預り、お釣り、お釣銭、釣銭、預り金、ポイント、残高、対象額、累計額は合計金額ではありません。
- 例: 合計 580 / お預り金 10,030 / お釣銭 9,450 の場合、amount は 580 です。
- 9,960円 のようなカンマ付き金額は 9960 として返してください。

【日付】
- date は「実際に購入・利用・診療・調剤した日」です。
- 最優先する日付は、レシート上部または明細上部にある「購入日時」「取引日時」「発行日」「領収日」「利用日」「診療日」「調剤日」です。
- レシートに時刻が併記されている日付を優先してください。例: 2026年06月07日(日) 13時40分。
- WORKMANなどで「返品・交換は購入日より14日以内」と書かれていても、その14日後の日付や期限はdateにしないでください。
- 「返品期限」「交換期限」「有効期限」「ポイント期限」「キャンペーン期限」「クーポン期限」「保証期限」「アプリ案内」「会員情報」「ポイント情報」「登録日」「印刷日」「締日」「請求日」はdateにしないでください。
- 電話番号、登録番号、伝票番号、会員番号、カード番号、バーコード番号を日付として使わないでください。
- 複数の日付がある場合は、一番上にある購入日時・取引日時・発行日・領収日を優先してください。
- 期限や案内文の日付しか読めない場合は、不明にしてください。
- 例: 購入日 2026-06-07、返品期限 2026-07-14 の場合、date は 2026-06-07 です。
- 例: レシート本文に 2026年06月07日(日) 13時40分 とある場合、date は 2026-06-07 です。
- 例: 下部のキャンペーン欄やポイント欄に 2023-10-01 があっても、購入日でなければ使わないでください。

【店名】
- 店名は領収書・レシートの発行元を読んでください。
- 「○○様」は宛名なので店名にしないでください。
- 人名を店名として出力しないでください。
- 病院・薬局の領収書では「医療機関」「薬局名」「発行元」の名称を店名にしてください。
- 店名・病院名は画像の文字をそのまま出力してください。
- 漢字を勝手に似た漢字へ補正しないでください。
- 例: 中岡内科 と書かれている場合は 中岡内科 と返してください。中田内科に変換してはいけません。

【区分】
- コンビニ、スーパー、飲食店は 食費。
- 電車、バス、高速代、駐車場、タクシーは 交通費。
- ホームセンター、WORKMAN、工具、作業用品は 工具費。
- 文房具、日用品、雑貨は 消耗品費。
- 携帯電話、ネット回線、通信料金は 通信費。
- 接待、会食、贈答は 交際費。
- 病院、クリニック、内科、歯科、薬局、調剤は 医療費。
- 判断できない場合は その他。

【その他】
- 読めない項目は 不明 または 0 にしてください。
- 推測で存在しない日付や店名を作らないでください。
"""

    response = client.chat.completions.create(
        model=os.environ.get("OPENAI_MODEL", "gpt-4o-mini"),
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            }
        ],
        temperature=0,
        response_format={"type": "json_object"},
    )

    print("========== OPENAI USAGE ==========")
    print(f"file: {image_path.name}")
    print(response.usage)
    print("==================================")

    save_usage_log(image_path.name, response.usage)

    content = response.choices[0].message.content

    try:
        data = json.loads(content)
    except Exception:
        data = {
            "date": "不明",
            "shop": "不明",
            "amount": 0,
            "category": "その他",
            "memo": "JSON解析失敗",
        }

    return {
        "date": str(data.get("date", "不明")),
        "shop": str(data.get("shop", "不明")),
        "amount": int(float(data.get("amount", 0) or 0)),
        "category": str(data.get("category", "その他")),
        "memo": str(data.get("memo", "")),
        "filename": image_path.name,
    }


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[col_letter].width = min(max(max_length + 3, 12), 40)


def create_excel(records: list[dict]) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "明細"

    headers = ["日付", "店名", "金額", "区分", "メモ", "ファイル名"]
    ws.append(headers)

    for record in records:
        ws.append([
            record["date"],
            record["shop"],
            record["amount"],
            record["category"],
            record["memo"],
            record["filename"],
        ])

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = '#,##0"円"'

    style_sheet(ws)

    summary_ws = wb.create_sheet("区分集計")
    summary_ws.append(["区分", "件数", "合計金額"])

    summary = {}
    for record in records:
        cat = record["category"]
        summary.setdefault(cat, {"count": 0, "amount": 0})
        summary[cat]["count"] += 1
        summary[cat]["amount"] += record["amount"]

    for cat, values in sorted(summary.items()):
        summary_ws.append([cat, values["count"], values["amount"]])

    for row in range(2, summary_ws.max_row + 1):
        summary_ws.cell(row=row, column=3).number_format = '#,##0"円"'

    style_sheet(summary_ws)

    total_ws = wb.create_sheet("合計")
    total_ws.append(["項目", "値"])
    total_ws.append(["件数", len(records)])
    total_ws.append(["合計金額", sum(r["amount"] for r in records)])
    total_ws.cell(row=3, column=2).number_format = '#,##0"円"'

    style_sheet(total_ws)

    output_path = OUTPUT_DIR / (
        f"receipt_summary_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_"
        f"{uuid4().hex[:8]}.xlsx"
    )
    wb.save(output_path)

    return output_path


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/analyze", methods=["POST"])
def analyze():
    if "receipts" not in request.files:
        flash("画像ファイルを選択してください。")
        return redirect(url_for("index"))

    files = request.files.getlist("receipts")

    if not files or files[0].filename == "":
        flash("画像ファイルを選択してください。")
        return redirect(url_for("index"))

    if len(files) > MAX_FILES_PER_UPLOAD:
        flash(f"一度にアップロードできるのは{MAX_FILES_PER_UPLOAD}枚までです。")
        return redirect(url_for("index"))

    new_records = []

    try:
        for f in files:
            if not allowed_file(f.filename):
                flash("対応形式は jpg / jpeg / png / webp / heic / heif です。")
                return redirect(url_for("index"))

            safe_name = secure_filename(f.filename)
            save_path = UPLOAD_DIR / (
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_"
                f"{uuid4().hex[:8]}_"
                f"{safe_name}"
            )

            converted_path = None

            try:
                f.save(save_path)

                converted_path = convert_heic_to_jpeg_if_needed(save_path)

                record = analyze_receipt(converted_path)
                new_records.append(record)

            finally:
                delete_file_safely(save_path)

                if converted_path and converted_path != save_path:
                    delete_file_safely(converted_path)

        accumulated_records = load_session_records()
        accumulated_records.extend(new_records)
        save_session_records(accumulated_records)

        excel_path = create_excel(accumulated_records)

    except Exception as e:
        flash(f"解析中にエラーが発生しました: {e}")
        return redirect(url_for("index"))

    return render_template(
        "result.html",
        records=accumulated_records,
        excel_filename=excel_path.name
    )


@app.route("/reset", methods=["POST", "GET"])
def reset():
    clear_session_records()
    flash("蓄積データをリセットしました。")
    return redirect(url_for("index"))


@app.route("/download/<filename>", methods=["GET"])
def download(filename):
    safe_name = secure_filename(filename)
    file_path = OUTPUT_DIR / safe_name

    if not file_path.exists():
        flash("Excelファイルが見つかりません。")
        return redirect(url_for("index"))

    return send_file(file_path, as_attachment=True)


@app.route("/usage-logs", methods=["GET"])
def usage_logs():
    logs = []
    total_count = 0
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_tokens = 0

    if supabase:
        try:
            logs_response = (
                supabase
                .table("usage_logs")
                .select("*")
                .order("id", desc=True)
                .limit(50)
                .execute()
            )

            logs = logs_response.data or []

            summary_response = (
                supabase
                .table("usage_logs")
                .select("prompt_tokens, completion_tokens, total_tokens")
                .execute()
            )

            all_logs = summary_response.data or []

            total_count = len(all_logs)
            total_prompt_tokens = sum(int(row.get("prompt_tokens") or 0) for row in all_logs)
            total_completion_tokens = sum(int(row.get("completion_tokens") or 0) for row in all_logs)
            total_tokens = sum(int(row.get("total_tokens") or 0) for row in all_logs)

        except Exception as e:
            print("Supabase read error:", e)

    total_cost_usd = estimate_cost_usd(total_prompt_tokens, total_completion_tokens)
    total_cost_jpy = estimate_cost_jpy(total_prompt_tokens, total_completion_tokens)

    html = f"""
    <!DOCTYPE html>
    <html lang="ja">
    <head>
        <meta charset="UTF-8">
        <title>Usage Logs</title>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <style>
            body {{
                font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
                background: #f5f8fd;
                padding: 24px;
                color: #0f172a;
            }}
            .container {{
                max-width: 1200px;
                margin: 0 auto;
                background: #fff;
                padding: 24px;
                border-radius: 16px;
                box-shadow: 0 8px 24px rgba(0,0,0,.08);
            }}
            h1 {{
                margin-top: 0;
            }}
            .small {{
                color: #64748b;
                font-size: 14px;
            }}
            .summary-grid {{
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 14px;
                margin: 22px 0;
            }}
            .summary-card {{
                background: #f8fafc;
                border: 1px solid #e5e7eb;
                border-radius: 14px;
                padding: 16px;
            }}
            .summary-label {{
                color: #64748b;
                font-size: 13px;
                margin-bottom: 6px;
            }}
            .summary-value {{
                font-size: 24px;
                font-weight: 900;
                color: #2563eb;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
            }}
            th, td {{
                border: 1px solid #e5e7eb;
                padding: 10px;
                font-size: 14px;
                text-align: left;
            }}
            th {{
                background: #dfeaff;
            }}
            @media(max-width:800px){{
                body {{
                    padding: 12px;
                }}
                .container {{
                    padding: 16px;
                    overflow-x: auto;
                }}
                .summary-grid {{
                    grid-template-columns: 1fr;
                }}
                table {{
                    min-width: 1100px;
                }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>Usage Logs</h1>
            <p class="small">
                Supabaseに保存された最新50件のOpenAI usageログです。<br>
                推定単価：入力 ${OPENAI_INPUT_PRICE_PER_1M} / 100万tokens、
                出力 ${OPENAI_OUTPUT_PRICE_PER_1M} / 100万tokens、
                為替 {USD_JPY_RATE}円/USD で計算。
            </p>

            <div class="summary-grid">
                <div class="summary-card">
                    <div class="summary-label">総件数</div>
                    <div class="summary-value">{total_count}</div>
                </div>
                <div class="summary-card">
                    <div class="summary-label">総トークン数</div>
                    <div class="summary-value">{total_tokens:,}</div>
                </div>
                <div class="summary-card">
                    <div class="summary-label">推定コスト USD</div>
                    <div class="summary-value">${total_cost_usd:.4f}</div>
                </div>
                <div class="summary-card">
                    <div class="summary-label">推定コスト 円</div>
                    <div class="summary-value">約{total_cost_jpy:.1f}円</div>
                </div>
            </div>

            <table>
                <tr>
                    <th>ID</th>
                    <th>日時</th>
                    <th>ファイル名</th>
                    <th>Prompt Tokens</th>
                    <th>Completion Tokens</th>
                    <th>Total Tokens</th>
                    <th>推定USD</th>
                    <th>推定円</th>
                    <th>IP</th>
                </tr>
    """

    for log in logs:
        prompt_tokens = int(log.get("prompt_tokens") or 0)
        completion_tokens = int(log.get("completion_tokens") or 0)
        total_tokens_row = int(log.get("total_tokens") or 0)
        cost_usd = estimate_cost_usd(prompt_tokens, completion_tokens)
        cost_jpy = estimate_cost_jpy(prompt_tokens, completion_tokens)

        html += f"""
                <tr>
                    <td>{log.get("id", "")}</td>
                    <td>{log.get("created_at", "")}</td>
                    <td>{log.get("filename", "")}</td>
                    <td>{prompt_tokens:,}</td>
                    <td>{completion_tokens:,}</td>
                    <td>{total_tokens_row:,}</td>
                    <td>${cost_usd:.6f}</td>
                    <td>約{cost_jpy:.2f}円</td>
                    <td>{log.get("ip_address", "")}</td>
                </tr>
        """

    html += """
            </table>
        </div>
    </body>
    </html>
    """

    return html


if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 5000)),
        debug=True
    )
