# -*- coding: utf-8 -*-
"""
領収書まとめ君 AI v3 Render版 - Windows PCアプリ
"""

from __future__ import annotations

import json
import os
import re
import sys
import threading
import traceback
from collections import defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_NAME = "領収書まとめ君 AI v3 Render版"
DEFAULT_SERVER_URL = "https://receipt-server-lnjo.onrender.com"
REQUEST_TIMEOUT_SEC = 120

IMAGE_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
    ".bmp",
    ".tif",
    ".tiff",
    ".heic",
    ".heif",
}

OUTPUT_COLUMNS = [
    "日付",
    "取引先",
    "店名",
    "金額",
    "区分",
    "メモ",
    "画像ファイル",
    "画像パス",
]

ENDPOINT_CANDIDATES = [
    "/analyze",
    "/analyze_receipt",
    "/analyze-receipt",
    "/receipt",
    "/receipts",
    "/extract",
    "/extract_receipt",
    "/api/analyze_receipt",
    "/api/receipt",
    "/api/extract",
]


@dataclass
class ReceiptRow:
    日付: str = ""
    取引先: str = ""
    店名: str = ""
    金額: int = 0
    区分: str = "未分類"
    メモ: str = ""
    画像ファイル: str = ""
    画像パス: str = ""


class ReceiptClientError(Exception):
    pass


class ReceiptServerClient:
    def __init__(self, base_url: str) -> None:
        self.base_url = self._normalize_base_url(base_url)
        self.working_endpoint: Optional[str] = None

    @staticmethod
    def _normalize_base_url(url: str) -> str:
        url = (url or "").strip()
        if not url:
            url = DEFAULT_SERVER_URL
        return url.rstrip("/")

    def health_check(self) -> Tuple[bool, str]:
        url = f"{self.base_url}/health"
        try:
            response = requests.get(url, timeout=30)
            text = response.text.strip()
            if response.ok:
                return True, text or "OK"
            return False, f"HTTP {response.status_code}: {text}"
        except Exception as exc:
            return False, str(exc)

    def analyze_image(self, image_path: Path) -> ReceiptRow:
        if not image_path.exists():
            raise ReceiptClientError(f"画像が見つかりません: {image_path}")

        endpoints = []

        if self.working_endpoint:
            endpoints.append(self.working_endpoint)

        endpoints.extend([ep for ep in ENDPOINT_CANDIDATES if ep not in endpoints])

        last_error = ""

        for endpoint in endpoints:
            try:
                data = self._post_image(endpoint, image_path)
                self.working_endpoint = endpoint
                return normalize_receipt_response(data, image_path)
            except Exception as exc:
                last_error = f"{endpoint}: {exc}"
                continue

        raise ReceiptClientError(
            "解析エンドポイントが見つかりませんでした。"
            "サーバー側のPOSTパスを確認してください。最後のエラー: "
            + last_error
        )

    def _post_image(self, endpoint: str, image_path: Path) -> Dict[str, Any]:
        url = f"{self.base_url}{endpoint}"
        mime_type = guess_mime_type(image_path)

        with image_path.open("rb") as f:
            files = {
                "file": (image_path.name, f, mime_type),
            }
            data = {
                "app": APP_NAME,
                "version": "v3-render-client",
                "mode": "receipt_excel",
            }
            response = requests.post(
                url,
                files=files,
                data=data,
                timeout=REQUEST_TIMEOUT_SEC,
            )

        if response.status_code in (404, 405):
            raise ReceiptClientError(f"HTTP {response.status_code}")

        if not response.ok:
            body = response.text[:1000]
            raise ReceiptClientError(f"HTTP {response.status_code}: {body}")

        try:
            return response.json()
        except Exception:
            text = response.text.strip()
            try:
                return json.loads(text)
            except Exception as exc:
                raise ReceiptClientError(f"JSONとして読めません: {text[:500]}") from exc


def guess_mime_type(path: Path) -> str:
    ext = path.suffix.lower()

    if ext in (".jpg", ".jpeg"):
        return "image/jpeg"
    if ext == ".png":
        return "image/png"
    if ext == ".webp":
        return "image/webp"
    if ext == ".bmp":
        return "image/bmp"
    if ext in (".tif", ".tiff"):
        return "image/tiff"
    if ext in (".heic", ".heif"):
        return "image/heic"

    return "application/octet-stream"


def first_value(data: Dict[str, Any], keys: List[str], default: Any = "") -> Any:
    for key in keys:
        if key in data and data[key] not in (None, ""):
            return data[key]
    return default


def flatten_response(data: Any) -> Dict[str, Any]:
    if isinstance(data, str):
        text = data.strip()
        try:
            loaded = json.loads(text)
            return flatten_response(loaded)
        except Exception:
            return {"メモ": text}

    if not isinstance(data, dict):
        return {"メモ": str(data)}

    for wrapper_key in ("result", "data", "receipt", "receipt_data", "analysis", "output"):
        value = data.get(wrapper_key)

        if isinstance(value, dict):
            return flatten_response(value)

        if isinstance(value, str):
            text = value.strip()
            try:
                loaded = json.loads(text)
                return flatten_response(loaded)
            except Exception:
                maybe_json = extract_json_object(text)
                if maybe_json:
                    return flatten_response(maybe_json)

    content = data.get("content") or data.get("message") or data.get("text")

    if isinstance(content, str):
        maybe_json = extract_json_object(content)
        if maybe_json:
            return flatten_response(maybe_json)

    return data


def extract_json_object(text: str) -> Optional[Dict[str, Any]]:
    text = text.strip()
    text = re.sub(r"^```(?:json)?", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"```$", "", text).strip()

    candidates = [text]
    start = text.find("{")
    end = text.rfind("}")

    if start >= 0 and end > start:
        candidates.append(text[start : end + 1])

    for candidate in candidates:
        try:
            loaded = json.loads(candidate)
            if isinstance(loaded, dict):
                return loaded
        except Exception:
            pass

    return None


def normalize_receipt_response(raw_data: Dict[str, Any], image_path: Path) -> ReceiptRow:
    data = flatten_response(raw_data)

    date_value = first_value(
        data,
        ["日付", "date", "transaction_date", "receipt_date", "発行日", "利用日", "購入日"],
        "",
    )

    partner = first_value(
        data,
        ["取引先", "partner", "vendor", "supplier", "company", "会社名", "宛先"],
        "",
    )

    store = first_value(
        data,
        ["店名", "store", "shop", "merchant", "store_name", "店舗名", "支払先"],
        "",
    )

    amount_value = first_value(
        data,
        ["金額", "amount", "total", "total_amount", "price", "税込金額", "合計", "総額"],
        0,
    )

    category = first_value(
        data,
        ["区分", "category", "expense_type", "勘定科目", "分類"],
        "未分類",
    )

    memo = first_value(
        data,
        ["メモ", "memo", "note", "notes", "description", "摘要", "内容"],
        "",
    )

    normalized_date = normalize_date(date_value)
    normalized_amount = normalize_amount(amount_value)

    if not partner and store:
        partner = store

    if not store and partner:
        store = partner

    return ReceiptRow(
        日付=normalized_date,
        取引先=str(partner).strip(),
        店名=str(store).strip(),
        金額=normalized_amount,
        区分=str(category or "未分類").strip() or "未分類",
        メモ=str(memo or "").strip(),
        画像ファイル=image_path.name,
        画像パス=str(image_path),
    )


def normalize_date(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    if not text:
        return ""

    text = text.replace("年", "/").replace("月", "/").replace("日", "")
    text = text.replace(".", "/").replace("-", "/")
    text = re.sub(r"\s+", "", text)

    era_match = re.search(r"令和\s*(\d+)\s*/\s*(\d+)\s*/\s*(\d+)", text)

    if era_match:
        y = 2018 + int(era_match.group(1))
        m = int(era_match.group(2))
        d = int(era_match.group(3))
        return f"{y:04d}-{m:02d}-{d:02d}"

    match = re.search(r"(20\d{2}|19\d{2})/(\d{1,2})/(\d{1,2})", text)

    if match:
        y, m, d = map(int, match.groups())
        return f"{y:04d}-{m:02d}-{d:02d}"

    match = re.search(r"(?<!\d)(\d{2})/(\d{1,2})/(\d{1,2})(?!\d)", text)

    if match:
        yy, m, d = map(int, match.groups())
        y = 2000 + yy
        return f"{y:04d}-{m:02d}-{d:02d}"

    match = re.search(r"(?<!\d)(\d{1,2})/(\d{1,2})(?!\d)", text)

    if match:
        y = datetime.now().year
        m, d = map(int, match.groups())

        if 1 <= m <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{m:02d}-{d:02d}"

    return text


def normalize_amount(value: Any) -> int:
    if value is None:
        return 0

    if isinstance(value, (int, float)):
        return int(round(float(value)))

    text = str(value).strip()

    if not text:
        return 0

    text = text.replace("円", "").replace("￥", "").replace("¥", "")
    text = text.replace(",", "")
    text = re.sub(r"[^0-9\.-]", "", text)

    if not text or text in ("-", ".", "-."):
        return 0

    try:
        return int(round(float(text)))
    except Exception:
        return 0


def month_key(date_text: str) -> str:
    text = (date_text or "").strip()

    match = re.match(r"^(\d{4})-(\d{2})-\d{2}$", text)

    if match:
        return f"{match.group(1)}-{match.group(2)}"

    match = re.match(r"^(\d{4})/(\d{1,2})/\d{1,2}$", text)

    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"

    return "日付不明"


def find_images_in_folder(folder: Path, recursive: bool = True) -> List[Path]:
    if not folder.exists() or not folder.is_dir():
        return []

    if recursive:
        paths = [p for p in folder.rglob("*") if p.is_file()]
    else:
        paths = [p for p in folder.iterdir() if p.is_file()]

    images = [p for p in paths if p.suffix.lower() in IMAGE_EXTENSIONS]
    images.sort(key=lambda p: str(p).lower())

    return images


def save_excel(rows: List[ReceiptRow], output_path: Path) -> None:
    wb = Workbook()

    ws_detail = wb.active
    ws_detail.title = "明細"

    ws_category = wb.create_sheet("区分集計")
    ws_month = wb.create_sheet("月別集計")

    write_detail_sheet(ws_detail, rows)
    write_category_sheet(ws_category, rows)
    write_month_sheet(ws_month, rows)

    wb.save(output_path)


def write_detail_sheet(ws, rows: List[ReceiptRow]) -> None:
    ws.append(OUTPUT_COLUMNS)

    for row in rows:
        d = asdict(row)
        ws.append([d.get(col, "") for col in OUTPUT_COLUMNS])

    apply_table_style(ws, header_row=1)

    set_widths(
        ws,
        {
            "A": 14,
            "B": 24,
            "C": 24,
            "D": 12,
            "E": 16,
            "F": 34,
            "G": 30,
            "H": 70,
        },
    )

    for row in range(2, ws.max_row + 1):
        ws[f"D{row}"].number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_category_sheet(ws, rows: List[ReceiptRow]) -> None:
    summary: Dict[str, Dict[str, int]] = defaultdict(lambda: {"件数": 0, "金額": 0})

    for row in rows:
        key = row.区分 or "未分類"
        summary[key]["件数"] += 1
        summary[key]["金額"] += int(row.金額 or 0)

    ws.append(["区分", "件数", "金額"])

    for key in sorted(summary.keys()):
        ws.append([key, summary[key]["件数"], summary[key]["金額"]])

    ws.append(["合計", len(rows), sum(int(r.金額 or 0) for r in rows)])

    apply_table_style(ws, header_row=1)
    set_widths(ws, {"A": 22, "B": 12, "C": 14})

    for row in range(2, ws.max_row + 1):
        ws[f"B{row}"].number_format = "#,##0"
        ws[f"C{row}"].number_format = "#,##0"

    ws.freeze_panes = "A2"


def write_month_sheet(ws, rows: List[ReceiptRow]) -> None:
    summary: Dict[str, Dict[str, int]] = defaultdict(lambda: {"件数": 0, "金額": 0})

    for row in rows:
        key = month_key(row.日付)
        summary[key]["件数"] += 1
        summary[key]["金額"] += int(row.金額 or 0)

    ws.append(["月", "件数", "金額"])

    for key in sorted(summary.keys()):
        ws.append([key, summary[key]["件数"], summary[key]["金額"]])

    ws.append(["合計", len(rows), sum(int(r.金額 or 0) for r in rows)])

    apply_table_style(ws, header_row=1)
    set_widths(ws, {"A": 16, "B": 12, "C": 14})

    for row in range(2, ws.max_row + 1):
        ws[f"B{row}"].number_format = "#,##0"
        ws[f"C{row}"].number_format = "#,##0"

    ws.freeze_panes = "A2"


def apply_table_style(ws, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if cell.row == header_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "合計":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF2CC")


def set_widths(ws, widths: Dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)

        if col_letter not in widths:
            ws.column_dimensions[col_letter].width = 16


class ReceiptApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()

        self.title(APP_NAME)
        self.geometry("1180x720")
        self.minsize(980, 620)

        self.image_paths: List[Path] = []
        self.rows: List[ReceiptRow] = []
        self.processing = False

        self.server_url_var = tk.StringVar(value=DEFAULT_SERVER_URL)
        self.folder_var = tk.StringVar(value="")
        self.output_var = tk.StringVar(value=str(Path.cwd() / "領収書まとめ_AI_v3.xlsx"))
        self.recursive_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="準備OK")
        self.progress_var = tk.DoubleVar(value=0)

        self._build_ui()

    def _build_ui(self) -> None:
        root = ttk.Frame(self, padding=10)
        root.pack(fill=tk.BOTH, expand=True)

        title = ttk.Label(root, text=APP_NAME, font=("Yu Gothic UI", 16, "bold"))
        title.pack(anchor=tk.W)

        server_frame = ttk.LabelFrame(root, text="Renderサーバー", padding=10)
        server_frame.pack(fill=tk.X, pady=(10, 6))
        server_frame.columnconfigure(1, weight=1)

        ttk.Label(server_frame, text="URL").grid(row=0, column=0, sticky=tk.W, padx=(0, 8))
        ttk.Entry(server_frame, textvariable=self.server_url_var).grid(row=0, column=1, sticky=tk.EW)
        ttk.Button(server_frame, text="/health確認", command=self.on_health_check).grid(row=0, column=2, padx=(8, 0))

        folder_frame = ttk.LabelFrame(root, text="画像フォルダ / Excel出力", padding=10)
        folder_frame.pack(fill=tk.X, pady=6)
        folder_frame.columnconfigure(1, weight=1)

        ttk.Label(folder_frame, text="画像フォルダ").grid(row=0, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        ttk.Entry(folder_frame, textvariable=self.folder_var).grid(row=0, column=1, sticky=tk.EW, pady=2)
        ttk.Button(folder_frame, text="選択", command=self.on_select_folder).grid(row=0, column=2, padx=(8, 0), pady=2)
        ttk.Button(folder_frame, text="画像読込", command=self.on_load_images).grid(row=0, column=3, padx=(8, 0), pady=2)

        ttk.Label(folder_frame, text="Excel保存先").grid(row=1, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        ttk.Entry(folder_frame, textvariable=self.output_var).grid(row=1, column=1, sticky=tk.EW, pady=2)
        ttk.Button(folder_frame, text="保存先選択", command=self.on_select_output).grid(row=1, column=2, padx=(8, 0), pady=2)
        ttk.Checkbutton(folder_frame, text="サブフォルダも対象", variable=self.recursive_var).grid(row=1, column=3, padx=(8, 0), pady=2)

        action_frame = ttk.Frame(root)
        action_frame.pack(fill=tk.X, pady=8)

        ttk.Button(action_frame, text="解析開始", command=self.on_start).pack(side=tk.LEFT)
        ttk.Button(action_frame, text="Excel保存", command=self.on_save_excel).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(action_frame, text="一覧クリア", command=self.on_clear).pack(side=tk.LEFT, padx=(8, 0))

        self.progress = ttk.Progressbar(action_frame, variable=self.progress_var, maximum=100)
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(16, 8))

        ttk.Label(action_frame, textvariable=self.status_var).pack(side=tk.RIGHT)

        table_frame = ttk.Frame(root)
        table_frame.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(table_frame, columns=OUTPUT_COLUMNS, show="headings")
        y_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)

        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        for col in OUTPUT_COLUMNS:
            self.tree.heading(col, text=col)

            width = {
                "日付": 100,
                "取引先": 160,
                "店名": 160,
                "金額": 90,
                "区分": 120,
                "メモ": 240,
                "画像ファイル": 220,
                "画像パス": 420,
            }.get(col, 140)

            anchor = tk.E if col == "金額" else tk.W
            self.tree.column(col, width=width, minwidth=70, anchor=anchor, stretch=True)

        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        note = ttk.Label(
            root,
            text="※ PC側ではOpenAI APIキーを使いません。画像はRenderサーバーへ送信され、結果だけExcel化します。",
            foreground="#555555",
        )
        note.pack(anchor=tk.W, pady=(6, 0))

    def on_health_check(self) -> None:
        client = ReceiptServerClient(self.server_url_var.get())
        ok, msg = client.health_check()

        if ok:
            messagebox.showinfo("health確認", f"OK\n\n{msg}")
            self.status_var.set("RenderサーバーOK")
        else:
            messagebox.showerror("health確認エラー", msg)
            self.status_var.set("Renderサーバー確認エラー")

    def on_select_folder(self) -> None:
        folder = filedialog.askdirectory(title="領収書画像フォルダを選択")

        if folder:
            self.folder_var.set(folder)
            default_output = Path(folder) / "領収書まとめ_AI_v3.xlsx"
            self.output_var.set(str(default_output))

    def on_select_output(self) -> None:
        initial = self.output_var.get().strip() or "領収書まとめ_AI_v3.xlsx"

        path = filedialog.asksaveasfilename(
            title="Excel保存先を選択",
            defaultextension=".xlsx",
            initialfile=Path(initial).name,
            filetypes=[("Excel files", "*.xlsx")],
        )

        if path:
            self.output_var.set(path)

    def on_load_images(self) -> None:
        folder_text = self.folder_var.get().strip()

        if not folder_text:
            messagebox.showwarning("確認", "画像フォルダを選択してください。")
            return

        folder = Path(folder_text)
        images = find_images_in_folder(folder, recursive=self.recursive_var.get())

        self.image_paths = images
        self.rows = []

        self.refresh_table()
        self.progress_var.set(0)
        self.status_var.set(f"画像 {len(images)} 件を読み込みました")

        if not images:
            messagebox.showinfo(
                "画像読込",
                "対象画像が見つかりませんでした。\n\n"
                "対応形式:\n"
                "jpg / jpeg / png / webp / bmp / tif / tiff / heic / heif",
            )

    def on_start(self) -> None:
        if self.processing:
            messagebox.showinfo("処理中", "現在解析中です。")
            return

        if not self.image_paths:
            self.on_load_images()

            if not self.image_paths:
                return

        self.processing = True
        self.rows = []

        self.refresh_table()
        self.progress_var.set(0)
        self.status_var.set("解析開始")

        thread = threading.Thread(target=self._process_images_thread, daemon=True)
        thread.start()

    def _process_images_thread(self) -> None:
        client = ReceiptServerClient(self.server_url_var.get())
        total = len(self.image_paths)

        try:
            for index, image_path in enumerate(self.image_paths, start=1):
                self.after(0, self.status_var.set, f"解析中 {index}/{total}: {image_path.name}")

                try:
                    row = client.analyze_image(image_path)
                except Exception as exc:
                    row = ReceiptRow(
                        日付="",
                        取引先="",
                        店名="",
                        金額=0,
                        区分="解析エラー",
                        メモ=str(exc),
                        画像ファイル=image_path.name,
                        画像パス=str(image_path),
                    )

                self.rows.append(row)
                self.after(0, self.add_row_to_table, row)
                self.after(0, self.progress_var.set, index / total * 100)

            output_path = Path(self.output_var.get().strip() or "領収書まとめ_AI_v3.xlsx")
            save_excel(self.rows, output_path)

            self.after(0, self.status_var.set, f"完了: {output_path}")
            self.after(
                0,
                messagebox.showinfo,
                "完了",
                f"解析とExcel保存が完了しました。\n\n{output_path}",
            )

        except Exception as exc:
            error_text = f"{exc}\n\n{traceback.format_exc()}"
            self.after(0, self.status_var.set, "エラー")
            self.after(0, messagebox.showerror, "エラー", error_text)

        finally:
            self.processing = False

    def on_save_excel(self) -> None:
        if not self.rows:
            messagebox.showwarning("確認", "保存する明細がありません。先に解析してください。")
            return

        output_path = Path(self.output_var.get().strip() or "領収書まとめ_AI_v3.xlsx")

        try:
            save_excel(self.rows, output_path)
            self.status_var.set(f"Excel保存完了: {output_path}")
            messagebox.showinfo("Excel保存", f"保存しました。\n\n{output_path}")

        except Exception as exc:
            messagebox.showerror("Excel保存エラー", str(exc))

    def on_clear(self) -> None:
        if self.processing:
            messagebox.showinfo("処理中", "解析中はクリアできません。")
            return

        self.rows = []
        self.image_paths = []

        self.refresh_table()
        self.progress_var.set(0)
        self.status_var.set("クリアしました")

    def refresh_table(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)

        for row in self.rows:
            self.add_row_to_table(row)

    def add_row_to_table(self, row: ReceiptRow) -> None:
        values = [getattr(row, col) for col in OUTPUT_COLUMNS]
        self.tree.insert("", tk.END, values=values)


def main() -> None:
    try:
        if getattr(sys, "frozen", False):
            os.chdir(Path(sys.executable).resolve().parent)
    except Exception:
        pass

    app = ReceiptApp()
    app.mainloop()


if __name__ == "__main__":
    main()
